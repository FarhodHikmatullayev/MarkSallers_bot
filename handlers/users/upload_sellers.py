import os

import psycopg2

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.types import ContentType, ContentTypes
from data.config import ADMINS, SUPERADMINS
from keyboards.default.back_menu import back_menu
from loader import dp, db
from openpyxl import load_workbook

from states.sellers import AddSellerState


# SAVE_DIRECTORY = 'path/to/save/directory'


@dp.message_handler(text='Upload sellers', user_id=SUPERADMINS, state='*')
async def get_sellers_excel_file(message: types.Message):
    text = "Sotuvchilar jadvalini yuboring (Excel)"
    await message.answer(text=text)
    await AddSellerState.waiting_for_excel_file.set()


@dp.message_handler(text='Upload sellers', state='*')
async def get_sellers_excel_file(message: types.Message):
    text = "Sizda bu buyruqdan foydalanish uchun ruxsat mavjud emas."
    await message.answer(text=text, reply_markup=back_menu)


@dp.message_handler(content_types=[ContentType.DOCUMENT], state=AddSellerState.waiting_for_excel_file)
async def save_sellers_from_excel(message: types.Message, state: FSMContext):
    if message.document.file_name.endswith('.xlsx') or message.document.file_name.endswith('.xls'):
        file_id = message.document.file_id
        file_name = message.document.file_name
        file_path = os.path.join('/tmp', file_name)

        await message.document.download(file_path)
        await db.delete_all_sellers()

        try:
            # Load the Excel file
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            # Insert data from the Excel file into the "employee" table
            for row in range(2, worksheet.max_row + 1):
                first_name = worksheet.cell(row=row, column=2).value
                last_name = worksheet.cell(row=row, column=3).value
                branch_name = worksheet.cell(row=row, column=4).value
                phone = worksheet.cell(row=row, column=5).value
                code = worksheet.cell(row=row, column=6).value
                print('branch_name', branch_name)
                branches = await db.select_branch(name=branch_name)
                if branches:
                    branch = branches[0]
                    branch_id = branch['id']

                else:
                    branch = await db.create_branch(name=branch_name)
                    branch_id = branch['id']
                seller = await db.create_seller(first_name=first_name, last_name=last_name, phone=phone,
                                                branch_id=branch_id, code=code)
                await state.finish()

            await message.answer(f'"{file_name}" dagi ma\'lumotlar muvaffaqiyatli saqlandi')
        except (Exception, psycopg2.Error) as error:
            await message.answer(f'Error occurred while processing the file: {error}\n'
                                 f'Kiritgan faylingizdagi ma\'lumotlarda xatolik mavjud,\n'
                                 f'Iltimos, o\'zgartirib qayta jo\'nating.')
            await AddSellerState.waiting_for_excel_file.set()
        finally:
            os.remove(file_path)

    else:
        await message.answer("Bu fayl excel fayl emas. Iltimos faqat excel fayl kiriting.")
        await AddSellerState.waiting_for_excel_file.set()


@dp.message_handler(state=AddSellerState.waiting_for_excel_file, content_types=ContentTypes.ANY)
async def save_employees_from_excel(message: types.Message):
    text = "Siz excel file kiritmadingiz, iltimos excel file kiriting"
    await message.answer(text=text)
    await AddSellerState.waiting_for_excel_file.set()
