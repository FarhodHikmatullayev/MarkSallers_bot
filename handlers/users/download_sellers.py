import tempfile
from typing import Union

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.types import CallbackQuery, Message
import openpyxl
import os

from data.config import ADMINS
from keyboards.default.back_menu import back_menu
from loader import dp, db, bot


async def download_all_sellers_function():
    sellers = await db.select_all_sellers()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = 'ISM'
    worksheet['C1'] = 'FAMILIYA'
    worksheet['D1'] = 'FILIAL'
    worksheet['E1'] = 'PHONE'
    worksheet['F1'] = 'ID'

    worksheet.cell(row=1, column=1, value='№')
    worksheet.cell(row=1, column=2, value='ISM')
    worksheet.cell(row=1, column=3, value="FAMILIYA")
    worksheet.cell(row=1, column=4, value='FILIAL')
    worksheet.cell(row=1, column=5, value='TELEFON RAQAM')
    worksheet.cell(row=1, column=6, value='ID')
    tr = 0
    for row, seller in enumerate(sellers, start=2):
        branch_id = seller['branch_id']

        branch = await db.select_branch(id=branch_id)

        branch_name = branch[0]['name']

        tr += 1
        worksheet.cell(row=row, column=1, value=tr)
        worksheet.cell(row=row, column=2, value=seller['first_name'])
        worksheet.cell(row=row, column=3, value=seller['last_name'])
        worksheet.cell(row=row, column=4, value=branch_name)
        worksheet.cell(row=row, column=5, value=seller['phone'])
        worksheet.cell(row=row, column=6, value=seller['code'])

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Sellers_data.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(text="Download sellers", user_id=ADMINS, state='*')
async def download_sellers(message: Message, state: FSMContext):
    await state.finish()
    temp_dir = await download_all_sellers_function()

    with open(os.path.join(temp_dir, 'Sellers_data.xlsx'), 'rb') as file:
        await message.answer_document(document=file)
        # await bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)

    os.remove(os.path.join(temp_dir, 'Sellers_data.xlsx'))


@dp.message_handler(text="Download sellers", state='*')
async def download_emp(message: Message, state: FSMContext):
    await state.finish()
    text = "Bu komanda faqat adminlar uchun"
    await message.answer(text=text, reply_markup=back_menu)
