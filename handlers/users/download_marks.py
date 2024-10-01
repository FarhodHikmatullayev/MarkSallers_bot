import datetime
import tempfile

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.types import Message
import openpyxl
import os

from data.config import ADMINS
from keyboards.default.back_menu import back_menu
from keyboards.inline.categories import categories_keyboard, categories_callback_data
from loader import dp, db, bot


async def download_all_comments_function(category_id):
    marks = await db.select_marks(category_id=category_id)
    category = await db.select_category(id=category_id)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = 'BAHOLAGAN SHAXS'
    worksheet['C1'] = 'FILIAL NOMI'
    worksheet['D1'] = 'XODIM ID RAQAMI'
    worksheet['E1'] = 'XODIM ISM FAMILIYASI'
    worksheet['F1'] = 'BAHOLASH KATEGORIYASI'
    worksheet['G1'] = 'BAHO'
    worksheet['H1'] = 'FIKR'
    worksheet['I1'] = 'VAQT'

    worksheet.cell(row=1, column=1, value='№')
    worksheet.cell(row=1, column=2, value='BAHOLAGAN SHAXS')
    worksheet.cell(row=1, column=3, value="FILIAL NOMI")
    worksheet.cell(row=1, column=4, value='XODIM ID RAQAMI')
    worksheet.cell(row=1, column=5, value='XODIM ISM FAMILIYASI')
    worksheet.cell(row=1, column=6, value='BAHOLASH KATEGORIYASI')
    worksheet.cell(row=1, column=7, value='BAHO')
    worksheet.cell(row=1, column=8, value='FIKR')
    worksheet.cell(row=1, column=9, value='VAQT')
    tr = 0
    for row, mark in enumerate(marks, start=2):
        user_id = mark['user_id']
        seller_id = mark['seller_id']
        users = await db.select_users(id=user_id)
        user = users[0]
        sellers = await db.select_sellers(id=seller_id)
        seller = sellers[0]
        seller_code = seller['code']
        seller_full_name = f"{seller['first_name']} {seller['last_name']}"
        branch_id = seller['branch_id']
        branch = await db.select_branch(id=branch_id)

        branch_name = branch[0]['name']
        full_name = user['full_name']
        username = user['username']
        phone = user['phone']
        telegram_id = user['telegram_id']

        tr += 1
        worksheet.cell(row=row, column=1, value=tr)
        worksheet.cell(row=row, column=2, value=full_name)
        worksheet.cell(row=row, column=3, value=branch_name)
        worksheet.cell(row=row, column=4, value=seller_code)
        worksheet.cell(row=row, column=5, value=seller_full_name)  # XODIM ISMI
        worksheet.cell(row=row, column=6, value=category['title'])
        worksheet.cell(row=row, column=7, value=mark['mark'])
        worksheet.cell(row=row, column=8, value=mark['description'])
        worksheet.cell(row=row, column=9,
                       value=(mark['created_at'] + datetime.timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S'))

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Comments_data.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(text='Bildirilgan baholarni yuklab olish', user_id=ADMINS, state='*')
async def download_mark(message: Message, state: FSMContext):
    await state.finish()

    categories = await db.select_all_categories()
    if categories:
        text = "Quyidagi baholash kategoriyalaridan birini tanlang"
        markup = await categories_keyboard()
        await message.answer(text=text, reply_markup=markup)
    else:
        text = "Hali fikrlar mavjud emas"
        await message.answer(text=text, reply_markup=back_menu)


@dp.callback_query_handler(categories_callback_data.filter())
async def get_category_for_download_marks(call: types.CallbackQuery, callback_data: dict):
    category_id = int(callback_data.get('category_id'))
    marks = await db.select_marks(category_id=category_id)
    if marks:
        temp_dir = await download_all_comments_function(category_id=category_id)

        with open(os.path.join(temp_dir, 'Comments_data.xlsx'), 'rb') as file:
            await call.message.answer_document(document=file)
            await bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)

        os.remove(os.path.join(temp_dir, 'Comments_data.xlsx'))
    else:
        text = 'Hali bu kategoriya bo\'yicha fikrlar bildirilmagan'
        await call.message.answer(text=text, reply_markup=back_menu)
        await bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)


@dp.message_handler(text="Bildirilgan baholarni yuklab olish", state='*')
async def download_emp(message: Message, state: FSMContext):
    await state.finish()
    text = "Bu komanda faqat adminlar uchun"
    await message.answer(text=text, reply_markup=back_menu)
